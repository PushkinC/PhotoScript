from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QLineEdit, QMainWindow
from PyQt5.QtGui import QIntValidator
from PyQt5.Qt import QPixmap
from Window_UI import Ui_MainWindow
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
from zipfile import ZipFile
from datetime import datetime
from time import sleep
import os, sys


class ResponseUser:
    def __init__(self, data):
        self.code = data[0]
        self.sale_cost = data[1]
        self.buy_cost = data[2]
        self.size = data[3]
        self.tp = data[4]
        self.sex = data[5]
        self.new_full_name = data[6]
        self.old_full_name = data[7]

class DataForSave():
    def __init__(self, data: ResponseUser, new_im:Image, old_im:Image):
        self.code = data.code
        self.sale_cost = data.sale_cost
        self.buy_cost = data.buy_cost
        self.size = data.size
        self.tp = data.tp
        self.sex = data.sex
        self.new_full_name = data.new_full_name
        self.old_full_name = data.old_full_name
        self.new_image = new_im
        self.old_image = old_im





class Window(QMainWindow, Ui_MainWindow):
    def __init__(self, dir_input='input', dir_output='output'):
        super().__init__()
        self.setupUi(self)
        self.dir_output = dir_output

        self.file_xlsx = load_workbook('book.xlsx')
        self.amount = len(os.listdir(dir_input))
        self.last_nam = self.getLast_nam()
        self.pictures = []
        for i, filename in enumerate(os.listdir(dir_input)):
            im = Image.open(f'input/{filename}')

            self.pictures.append({
                'old_path': dir_input,
                'new_path': dir_output,
                'old_name': filename,
                'new_name': f'{i + 1 + self.last_nam}.jpg',
                'image': im,
                'size': im.size,
                'h': im.height,
                'w': im.width
            })

        if len(self.pictures) == 0:
            print('Нету фото')
            sleep(5)
            sys.exit()

        self.cur_nam = 0
        self.lst_for_del = []
        self.lst_of_data = [None] * self.amount

        self.initUI()

    def initUI(self):
        self.setWindowTitle('Фото')
        self.counter.setText(f'{self.cur_nam + 1} из {self.amount}')
        self.load_comboBox()

        self.btn_back.clicked.connect(self.cl_back)
        self.btn_next.clicked.connect(self.cl_next)
        self.btn_back.setEnabled(False)

        self.lbl_name_old.setText(f'Имя оригинала: {self.pictures[self.cur_nam]["old_name"]}')
        self.lbl_name_new.setText(f'Новое имя: {self.pictures[self.cur_nam]["new_name"]}')

        rx = QIntValidator()
        self.line_cost_sale.setValidator(rx)
        self.line_cost_buy.setValidator(rx)

        self.label.setPixmap(QPixmap(self.pictures[self.cur_nam]['old_path'] + '/' + self.pictures[self.cur_nam]['old_name']).scaled(self.label.width(), self.label.height(), QtCore.Qt.AspectRatioMode.KeepAspectRatio))

    def load_comboBox(self):
        with open('comboboxitems/sex.txt', 'rt', encoding="UTF-8") as f:
            data = []
            for i in f.readlines():
                data.append(i.rstrip('\n'))
        self.sex.clear()
        self.sex.addItems(data)

        with open('comboboxitems/types of clothes.txt', 'rt', encoding="UTF-8") as f:
            data = []
            for i in f.readlines():
                data.append(i.rstrip('\n'))
        self.tp.clear()
        self.tp.addItems(data)

    def cl_next(self):
        if self.line_size.text() and self.line_cost_sale.text() and self.line_cost_buy.text():
            data = self.getData()
            # self.save_to_excel(data)
            self.lst_of_data[self.cur_nam] = self.createIm(data)
        else:
            return

        self.cur_nam += 1
        self.btn_back.setEnabled(True)
        if self.cur_nam + 1 == self.amount:
            self.btn_next.setText('Готово')
        elif self.cur_nam >= self.amount:
            self.cur_nam -= 1
            self.save_all()

        self.changeScreen()

    def cl_back(self):
        self.cur_nam -= 1
        if self.cur_nam <= 0:
            self.btn_back.setEnabled(False)
            self.cur_nam = 0
        self.changeScreen()


    def changeScreen(self):
        self.lbl_name_old.setText(f'Имя оригинала: {self.pictures[self.cur_nam]["old_name"]}')
        self.lbl_name_new.setText(f'Новое имя: {self.pictures[self.cur_nam]["new_name"]}')
        self.label.setPixmap(QPixmap(self.pictures[self.cur_nam]['old_path'] + '/' + self.pictures[self.cur_nam]['old_name']).scaled(self.label.width(), self.label.height(), QtCore.Qt.AspectRatioMode.KeepAspectRatio))
        self.counter.setText(f'{self.cur_nam + 1} из {self.amount}')

        if self.lst_of_data[self.cur_nam] is None:
            self.line_size.setText('')
            self.line_cost_sale.setText('')
            self.line_cost_buy.setText('')
            self.tp.setCurrentIndex(0)
            self.sex.setCurrentIndex(0)
        else:
            self.line_size.setText(str(self.lst_of_data[self.cur_nam].size))
            self.line_cost_sale.setText(str(self.lst_of_data[self.cur_nam].sale_cost))
            self.line_cost_buy.setText(str(self.lst_of_data[self.cur_nam].buy_cost))
            index = self.tp.findText(self.lst_of_data[self.cur_nam].tp)
            self.tp.setCurrentIndex(index)
            index = self.sex.findText(self.lst_of_data[self.cur_nam].sex)
            self.sex.setCurrentIndex(index)

    def getLast_nam(self):
        last_nam = 0
        for i in os.listdir(self.dir_output):
            if int(i.rstrip('.jpg')) >= last_nam:
                last_nam = int(i.rstrip('.jpg'))
        return last_nam

    def getData(self) -> ResponseUser:
        data = [
            self.cur_nam + self.last_nam + 1,
            int(self.line_cost_sale.text()),
            int(self.line_cost_buy.text()),
            self.line_size.text(),
            self.tp.currentText(),
            self.sex.currentText(),
            self.pictures[self.cur_nam]['new_path'] + '/' +  self.pictures[self.cur_nam]['new_name'],
            self.pictures[self.cur_nam]['old_path'] + '/' +  self.pictures[self.cur_nam]['old_name']
        ]
        return ResponseUser(data)

    def save_to_excel(self, data):
        sheets = self.file_xlsx.sheetnames
        sheet_name = sheets[0]
        sheet = self.file_xlsx[sheet_name]
        max_row = sheet.max_row + 1
        sheet[f'A{max_row}'].value = data.code
        sheet[f'B{max_row}'].value = data.tp
        sheet[f'C{max_row}'].value = data.sex
        sheet[f'D{max_row}'].value = data.size
        sheet[f'E{max_row}'].hyperlink = data.new_full_name
        sheet[f'E{max_row}'].value = data.new_full_name
        sheet[f'E{max_row}'].style = "Hyperlink"
        sheet[f'F{max_row}'].value = data.sale_cost
        sheet[f'G{max_row}'].value = data.buy_cost
        sheet[f'H{max_row}'].value = f'=F{max_row} - G{max_row}'
        sheet[f'I{max_row}'].value = data.old_full_name.split('/')[-1]
        sheet[f'J{max_row}'].value = datetime.now().date()

        self.file_xlsx.save('book.xlsx')

    def createIm(self, data) -> DataForSave:
        im = self.pictures[self.cur_nam]['image']
        im_size = im.size
        new_im = Image.new('RGB', (im_size[0], im_size[1] + 400), color=(255, 255, 255))
        new_im.paste(im)
        new_size = new_im.size

        font = ImageFont.truetype('arial.ttf', size=200)
        text = f'{data.sale_cost} руб.,  Размер {data.size},  Арт. {self.cur_nam + 1 + self.last_nam}'

        draw_im = ImageDraw.Draw(new_im)
        _, _, w, h = draw_im.textbbox((0, 0), text, font=font)

        draw_im.text(((new_size[0] - w) // 2, new_size[1] - h - 50), text, (0, 0, 0), font=font)


        data = DataForSave(data, new_im, im)
        return data


    def save_all(self):
        with ZipFile('archive.zip', 'a') as f:
            for data in self.lst_of_data:
                f.write(f'{data.old_full_name}', f'Save from {datetime.now().date()}/{data.old_full_name.split("/")[-1]}')
                data.new_image.save(data.new_full_name)
                self.save_to_excel(data)
                os.remove(f'{data.old_full_name}')

        sys.exit()


