from PIL import Image, ImageDraw, ImageFont, ImageTk
from tkinter import Label, Tk, Entry
from openpyxl import load_workbook
from window import Window
from PyQt5.QtWidgets import QApplication
import os, sys

def except_hook(cls, exception, traceback):
    sys.__excepthook__(cls, exception, traceback)

file_xlsx = load_workbook('book.xlsx')

def save_to_excel(data):
    global file_xlsx
    sheets = file_xlsx.sheetnames
    sheet_name = sheets[0]
    sheet = file_xlsx[sheet_name]
    max_row = sheet.max_row + 1
    sheet[f'A{max_row}'].value = data[0]
    sheet[f'B{max_row}'].value = data[1]
    sheet[f'C{max_row}'].value = data[2]
    sheet[f'D{max_row}'].value = data[3]
    sheet[f'E{max_row}'].value = f'=B{max_row} - C{max_row}'
    sheet[f'F{max_row}'].hyperlink = data[4]
    sheet[f'F{max_row}'].value = data[4]
    sheet[f'F{max_row}'].style = "Hyperlink"



app = QApplication(sys.argv)
ex = Window('input', 'output')
ex.show()
sys.excepthook = except_hook
sys.exit(app.exec_())


file_xlsx.save('book.xlsx')
file_xlsx.close()



